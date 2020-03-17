import openpyxl
from openpyxl import load_workbook
import pandas as pd
import numpy as np
wb = openpyxl.load_workbook("create_sample.xlsx")
nws=wb.create_sheet(index=0,title="Sheet3")
ws=wb.save('create_sample.xlsx')
#wa=wb.sheetnames
#print(wa)
c=nws['B5']
print(c.value)



# wb = load_workbook('I:\2020 Python 強攻\Financial Sample.xlsx')
# print(wb.sheetnames)
#from openpyxl import Workbook
# 創建一個空白活頁簿物件
#wb = Workbook()
# 選取正在工作中的表單
#ws = wb.active
# 指定值給 A1 儲存格
#ws['A1'] = '我是儲存格'
# 向下新增一列並連續插入值
#ws.append([1, 2, 3])
#ws.append([3, 2, 1])
# 儲存成 create_sample.xlsx 檔案
#wb.save('create_sample.xlsx')
#=============================
